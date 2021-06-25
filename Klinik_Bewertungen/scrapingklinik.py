from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
# from selenium.webdriver.chrome.options import Options
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl

kl_name   = []
text_bew  = []
datum_bew = []
stern_bew = []
like_bew  = []

wb = openpyxl.load_workbook('Klinikliste.xlsx')
sheet = wb['Tabelle1']
url = [] # ist eine liste von allen Kliniken Links

def new_func(sheet, url):
    for row in sheet.iter_rows(
    min_row= 2,
    max_row=16,
    min_col=2,
    max_col=2,
    values_only=True):
     url.append(str(row))
     for s in url:
         s.replace(",",'')  

new_func(sheet, url)


     
print(type(url))
print(type(url[0]))
print(url) 
# print(len(url))  
# print(li)     
#  for i in url:
#      i=",".join(i)

#print(url)      
for link in str(url):      
    while True:
        
        lastlink=0
        
        DRIVER_PATH = '../GoogleMaps_Scraping/chromedriver'
        options = webdriver.ChromeOptions()
        
        #chromedriver language change to German
        options.add_experimental_option('prefs', {'intl.accept_languages': 'de'})
        driver =driver = webdriver.Chrome(ChromeDriverManager().install()) #webdriver.Chrome(executable_path=DRIVER_PATH,options=options)
        
        driver.get(url)
        driver.maximize_window()
        time.sleep(3)
        
        #Cookies
        driver.find_element_by_xpath('/html/body/c-wiz/div/div/div/div[2]/div[1]/div[4]/form/div[1]/div/button/span').click()
        time.sleep(4)

        klinikName = driver.find_element_by_id("searchboxinput").get_attribute("value")

        #bewertungen clicken
        driver.find_element_by_css_selector('.widget-pane-link').click()
        time.sleep(3)

        #scroll
        jscommand = """
        berichte = document.querySelector(".section-layout.section-scrollbox.cYB2Ge-oHo7ed.cYB2Ge-ti6hGc");
        berichte.scrollTo(0, berichte.scrollHeight);var lenOfPage=berichte.scrollHeight;return lenOfPage;
        """
        lenOfPage = driver.execute_script(jscommand)
        match=False
        while(match==False):
            lastCount = lenOfPage
            time.sleep(1)
            lenOfPage = driver.execute_script(jscommand)
            if lastCount == lenOfPage:
                match=True
        time.sleep(1)
        
        #mehr button click
        loads = driver.find_elements_by_css_selector('.ODSEW-KoToPc-ShBeI.gXqMYb-hSRGPd')
        for load in loads:
            load.click()
        time.sleep(1)
        
        texts = driver.find_elements_by_xpath('//div[@class="ODSEW-ShBeI-ShBeI-content"]/span[2]')
        dates = driver.find_elements_by_xpath('//span[@class="ODSEW-ShBeI-RgZmSc-date"]')
        stars = driver.find_elements_by_xpath('//div[@class="ODSEW-ShBeI-jfdpUb"]/span[2]')
        likes = driver.find_elements_by_xpath('//button[@class="ODSEW-ShBeI-Sc2xXc-LgbsSe"]/span/span[2]')

        klinikName_list = [klinikName for a in range(len(texts))]
        text_list = [a.text for a in texts]
        date_list = [a.text for a in dates]
        star_list = [a.get_attribute("aria-label")[1] for a in stars]

        likes_list =[]
        for t in likes:
            if t.text =='':
                likes_list.append('Keine')
            else:
                likes_list.append(t.text)

        like_list=[] 
        for t in range(len(texts)):
            if t>=len(likes_list):
                like_list.append('Keine')
            else:
                like_list.append(likes_list[t])
                
        for i in klinikName_list:
            kl_name.append(i)
        for i in text_list:
            text_bew.append(i)
        for i in date_list:
            datum_bew.append(i)
        for i in star_list:
            stern_bew.append(i)
        for i in like_list:
            like_bew.append(i)    
        
        driver.quit() 
        
        lastlink +=1 
        
        if (lastlink < len(url)):
            print("es wurden", len(url), "klinken gesucht")
            break
            
                

            
        
# df = pd.DataFrame(zip(kl_name,text_bew,datum_bew,stern_bew,like_bew), columns=["Name der Klinik","Bewertung","Datum der Bewertung","Sternebewertung",'Likes'])
# df.to_csv('GoogleMaps111.csv', index=False)   



