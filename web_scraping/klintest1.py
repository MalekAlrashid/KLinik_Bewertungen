
import requests
from bs4 import BeautifulSoup
import csv
from itertools import zip_longest
import selenium
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
import openpyxl
from openpyxl import load_workbook
wb = openpyxl.load_workbook('Klinikliste.xlsx')
sheet = wb['Tabelle1']
url = [] # ist eine liste von allen Kliniken Links
for row in sheet.iter_rows(
     min_row= 2,
     max_row=16,
     min_col=3,
     values_only=True):

     url.append(row)
     
klinik_name = []
titel = []
datum_bew = []
fachbereich = []
gesamtzufriedenheit = []
qualität_der_Beratung = []
mediz_Behandlung = []
verwaltung_Abläufe = []
ausstattung_Gestaltung = []
erfahrungsbericht = []
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(ChromeDriverManager().install())

result =  requests.get("https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-kliniken-herzogin-elisabeth-braunschweig")
src = result.content
soup = BeautifulSoup(src, 'html.parser')
for element in soup.find("h1", attrs={"itemprop":"name"}):
    klinik_name.append(element.text)
    
for elem in soup.find_all("article", attrs={"class":"bewertung"}):
    titel_ = elem.find_all("header > h2", attrs={"itemprop":"name"})
    titel.append(titel_.text)
    time = elem.find_all("time")
    datum_bew.append(time.text)
    fach_bereich = elem.find.all("a", attrs={"class":"js-tooltip"})
    fachbereich.append(fach_bereich.text)
    
for st in soup.find_all("section > dd"):
    gz = soup.find("dt")
    gesamtzufriedenheit.append(gz.text)
    qb = soup.find("dt")
    qualität_der_Beratung.append(qb.text)
    mb = soup.find("dt")
    mediz_Behandlung.append(mb.text)
    vua = soup.find("dt")
    verwaltung_Abläufe.append(vua.text)
        
print(gesamtzufriedenheit)        
            
        
    
    
    
    


# soup = BeautifulSoup(src, "lxml")



# kl_name = soup.find('h1', attrs={'style':"display:inline-block"})
# bew_text = soup.find_all('p', attrs={'itemprop':'reviewBody'})
# stern_bew = soup.find_all('img', attrs={'class':'star-6'})
# datum = soup.find_all('time', attrs={'pubdate':"pubdate"})
# kl_link = soup.find('a', attrs={'class':"icon16 icon-globe-fill"})

# for i in range(len(kl_name)):
#     klinik_name.append(kl_name.text)
# for i in range(len(bew_text)):
#     bewerbungs_text.append(bew_text[i].text)
# for i in range(len(stern_bew)):
#     stern_bewerbung.append(stern_bew[i].text)
# for i in range(len(datum)):
#    datum_bew.append(datum[i].text)
# for i in range(len(kl_link)):
#        klinik_link.append(kl_link)


# with open("C:/Users/Alras/kl11.csv", "w",encoding='UTF8', newline='') as myfile:
#     header = ['Klinik Name', 'Bewertung', 'Stern Bewertung', 'Datum' , 'Link der Klinikum']
#     wr = csv.writer(myfile)
#     wr.writerow(header)

# for i in url:
#     print(">>>>>  ",i, "  <<<<")

# print



